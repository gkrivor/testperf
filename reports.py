import sys
import subprocess
import datetime
from copy import deepcopy

def performance_report(model,model_name, read_times, inference_times, warm_up_times, batches):
  try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference, Series
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    main_sheet = wb.active
    read_sheet = wb.create_sheet("Read")
    inference_sheet = wb.create_sheet("Inference")

    read_sheet.append(["Reading times"])

    offset_col = 2
    offset_stat_row = read_sheet.max_row + 1
    read_sheet.append(["Average"])
    read_sheet.append(["Median"])
    read_sheet.append(["90th Percentile"])
    read_sheet.append(["95th Percentile"])
    read_sheet.append(["99th Percentile"])
    read_sheet.append(["Minimum"])
    read_sheet.append(["Maximum"])

    col_letter = get_column_letter(offset_col)
    offset_row = read_sheet.max_row + 1
    last_row = offset_row + len(read_times) - 2
    read_sheet[col_letter + str(offset_stat_row + 0)] = "=AVERAGE(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
    read_sheet[col_letter + str(offset_stat_row + 1)] = "=MEDIAN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
    read_sheet[col_letter + str(offset_stat_row + 2)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.9)"
    read_sheet[col_letter + str(offset_stat_row + 3)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.95)"
    read_sheet[col_letter + str(offset_stat_row + 4)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.99)"
    read_sheet[col_letter + str(offset_stat_row + 5)] = "=MIN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
    read_sheet[col_letter + str(offset_stat_row + 6)] = "=MAX(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"

    read_sheet.append(["Run", "Time (s)"])

    offset_col = 2
    offset_row = read_sheet.max_row + 1
    last_row = offset_row + len(read_times) - 2

    idx = 1
    for item in read_times[:-1]:  
        read_sheet.append([idx, item])
        idx += 1

    series = Series(values=Reference(read_sheet, min_col=offset_col, min_row=offset_row, max_col=offset_col, max_row=last_row), title="Reading times")
    if len(read_times) <= 2:
        series.marker.symbol = "circle"
        series.marker.size = 6
    chart = LineChart()
    chart.series.append(series)
    chart.title = "Reading times"
    chart.x_axis.title = "Run"
    chart.y_axis.title = "Time (s)"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend = None
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.75, w=0.9,
        )
    )
    read_sheet.add_chart(chart, "C1")
    main_sheet.add_chart(deepcopy(chart), "D35")

    inference_sheet.append(["Inference times"])
    inference_table = [[] for _ in range(len(batches))]
    x_axis_max = 0
    x_axis_min = float('inf')
    for batch_index in range(len(batches)):
        for item in inference_times[batches[batch_index]][:-1]:
            inference_table[batch_index].append(item)
            x_axis_max = max(x_axis_max, item)
            x_axis_min = min(x_axis_min, item)

    # Table header
    inference_sheet.append(["Metric"] + [f"Batch {batch}" for batch in batches])
    # Aggregated statistics
    inference_sheet.column_dimensions[get_column_letter(1)].width = 30
    offset_col = 2
    offset_stat_row = inference_sheet.max_row + 1
    inference_sheet.append(["Average"])
    inference_sheet.append(["Median"])
    inference_sheet.append(["90th Percentile"])
    inference_sheet.append(["95th Percentile"])
    inference_sheet.append(["99th Percentile"])
    inference_sheet.append(["Minimum"])
    inference_sheet.append(["Maximum"])
    inference_sheet.append(["IPS (Average)"])
    inference_sheet.append(["IPS (Median)"])
    inference_sheet.append(["IPS (90th Percentile)"])
    inference_sheet.append(["IPS (95th Percentile)"])
    inference_sheet.append(["IPS (99th Percentile)"])
    inference_sheet.append(["BPS (Average)"])
    inference_sheet.append(["BPS (Median)"])
    inference_sheet.append(["BPS (90th Percentile)"])
    inference_sheet.append(["BPS (95th Percentile)"])
    inference_sheet.append(["BPS (99th Percentile)"])
    inference_sheet.append(["Warm Up Time"])

    # Table header
    inference_sheet.append(["Run"] + [f"Batch {batch}" for batch in batches])

    offset_col = 2
    offset_row = inference_sheet.max_row + 1
    last_row = offset_row + len(inference_table[0]) - 2

    for batch_index in range(len(batches)):
        col_letter = get_column_letter(offset_col + batch_index)
        inference_sheet[col_letter + str(offset_stat_row + 0)] = "=AVERAGE(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        inference_sheet[col_letter + str(offset_stat_row + 1)] = "=MEDIAN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        inference_sheet[col_letter + str(offset_stat_row + 2)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.9)"
        inference_sheet[col_letter + str(offset_stat_row + 3)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.95)"
        inference_sheet[col_letter + str(offset_stat_row + 4)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.99)"
        inference_sheet[col_letter + str(offset_stat_row + 5)] = "=MIN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        inference_sheet[col_letter + str(offset_stat_row + 6)] = "=MAX(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        # Inference Per Second depending on calculated time
        inference_sheet[col_letter + str(offset_stat_row + 7)] = "=1000 / " + col_letter + str(offset_stat_row + 0)
        inference_sheet[col_letter + str(offset_stat_row + 8)] = "=1000 / " + col_letter + str(offset_stat_row + 1)
        inference_sheet[col_letter + str(offset_stat_row + 9)] = "=1000 / " + col_letter + str(offset_stat_row + 2)
        inference_sheet[col_letter + str(offset_stat_row + 10)] = "=1000 / " + col_letter + str(offset_stat_row + 3)
        inference_sheet[col_letter + str(offset_stat_row + 11)] = "=1000 / " + col_letter + str(offset_stat_row + 4)
        # Batch Per Second depending on IPS
        inference_sheet[col_letter + str(offset_stat_row + 12)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 7)
        inference_sheet[col_letter + str(offset_stat_row + 13)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 8)
        inference_sheet[col_letter + str(offset_stat_row + 14)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 9)
        inference_sheet[col_letter + str(offset_stat_row + 15)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 10)
        inference_sheet[col_letter + str(offset_stat_row + 16)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 11)
        inference_sheet[col_letter + str(offset_stat_row + 17)] = str(warm_up_times[batches[batch_index]])

    chart = LineChart()
    chart.title = "Metrics"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Time (s)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = x_axis_max
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile", "Minimum", "Maximum"]
    for metric_index in range(len(metrics)):
        series = Series(values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + metric_index, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + metric_index), title=f"{metrics[metric_index]}")
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
    batch_titles = Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row - 1, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row - 1)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    chart.width = 25
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "1")
    main_sheet.add_chart(deepcopy(chart), "D5")

    chart = LineChart()
    chart.title = "IPS"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Inferences Per Second"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile"]
    for metric_index in range(len(metrics)):
        series = Series(values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + 7 + metric_index, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + 7 + metric_index), title=f"{metrics[metric_index]}")
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
    batch_titles = Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row - 1, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row - 1)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    chart.width = 15
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "16")
    main_sheet.add_chart(deepcopy(chart), "D20")

    chart = LineChart()
    chart.title = "BPS"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Batches Per Second"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile"]
    for metric_index in range(len(metrics)):
        series = Series(values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + 12 + metric_index, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + 12 + metric_index), title=f"{metrics[metric_index]}")
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
    batch_titles = Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row - 1, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row - 1)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    chart.width = 15
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 11) + "16")
    main_sheet.add_chart(deepcopy(chart), "N20")

    idx = 0
    for idx in range(1, len(inference_table[0])):
        row = [idx]
        for batch_index in range(len(batches)):
            row.append(inference_table[batch_index][idx - 1] if len(inference_table[batch_index]) > idx - 1 else None)
        inference_sheet.append(row)
    chart = LineChart()
    chart.title = "Inference times"
    chart.x_axis.title = "Run"
    chart.y_axis.title = "Time (s)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = x_axis_max
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    for batch_index in range(len(batches)):
        series = Series(values=Reference(inference_sheet, min_col=batch_index + offset_col, min_row=offset_row, max_col=batch_index + offset_col, max_row=last_row), title=f"Batch {batches[batch_index]}")
        chart.series.append(series)
    chart.width = 15
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "33")
    main_sheet.add_chart(deepcopy(chart), "N35")

    report_datetime = datetime.datetime.now()
    main_sheet.title = "Overview"
    main_sheet.column_dimensions[get_column_letter(1)].width = 30
    main_sheet.append(['Model:', model_name])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Description:', str(model)])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Run Command:', ' '.join(sys.argv)])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Report Date:', report_datetime.strftime('%Y-%m-%d %H:%M:%S')])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=6)
    main_sheet.append(['Batches:', *batches])
    main_sheet.append(['Total Inference Runs:', model.total_inference_runs])
    main_sheet.append([])
    main_sheet.append(['System Information:'])
    try:
        import platform
        main_sheet.append(['Hostname:', platform.node()])
        main_sheet.append(['OS:', platform.system()])
        main_sheet.append(['OS Version:', platform.version()])
        main_sheet.append(['OS Release:', platform.release()])
    except Exception as e:
        main_sheet.append([f'Cannot get OS information {e}'])
    main_sheet.append(['Python Version:', sys.version])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)

    try:
        result = subprocess.run(
            ['pip', 'list', '--format', 'columns'],
            capture_output=True,
            text=True
        )
        output = result.stdout.split('\n')
        for item in output:
            main_sheet.append(item.split())
    except Exception as e:
        main_sheet.append([f'Cannot get PIP list {e}'])

    workbook_path = f"{platform.node().lower()}_{model_name}_{report_datetime.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(workbook_path)
    print(f"{{ \"Workbook\": \"{workbook_path}\" }}")

  except Exception as e:
    print(f'{{ "Error": "Failed to load openpyxl {e}" }}')
  return workbook_path
